import nltk
from nltk.tokenize import word_tokenize
from flask import Flask, render_template, request, jsonify, session
import json
import os
import openpyxl
from rapidfuzz import process, fuzz
from nltk.corpus import wordnet
import requests
import numpy as nk
import random
#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

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
# Ensure required NLTK data is available (download if missing)
try:
    nltk.data.find("corpora/wordnet")
except LookupError:
    nltk.download("wordnet")
    nltk.download("omw-1.4")

try:
    nltk.data.find("tokenizers/punkt")
except LookupError:
    nltk.download("punkt")

# Load FAQs safely
FAQS = []
faq_path = os.path.join(os.path.dirname(__file__), "admission_faq.json")
try:
    with open(faq_path, "r", encoding="utf-8") as f:
        FAQS = json.load(f)
except FileNotFoundError:
    print(f"Warning: FAQ file not found at {faq_path}. Continuing with empty FAQ list.")
except json.JSONDecodeError as e:
    print(f"Warning: Could not parse FAQ file {faq_path}: {e}. Continuing with empty FAQ list.")

@app.route("/admission_enquiry", methods=["POST"])
def admission_enquiry():
    data = request.get_json()
    name = data.get("name", "")
    email = data.get("email", "")
    phone = data.get("phone", "")
    course = data.get("course", "")
    enquiry = {"name": name, "email": email, "phone": phone, "course": course}
    try:
        excel_path = "admission_enquiries.xlsx"
        if not os.path.exists(excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Admission Enquiries"
            ws.append(["Name", "Email", "Phone", "Course"])
            wb.save(excel_path)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        ws.append([name, email, phone, course])
        wb.save(excel_path)
        return jsonify({"success": True, "message": "Admission enquiry submitted successfully!", "details": enquiry})
    except Exception as e:
        return jsonify({"success": False, "message": f"Failed to submit admission enquiry: {e}"}), 500

@app.route("/book_callback", methods=["POST"])
def book_callback():
    data = request.get_json()
    name = data.get("name", "")
    phone = data.get("phone", "")
    preferred_time = data.get("preferred_time", "")
    callback = {"name": name, "phone": phone, "preferred_time": preferred_time}
    try:
        try:
            with open("callbacks.json", "r", encoding="utf-8") as f:
                callbacks = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            callbacks = []
        callbacks.append(callback)
        with open("callbacks.json", "w", encoding="utf-8") as f:
            json.dump(callbacks, f, ensure_ascii=False, indent=2)

        excel_path = "callbacks.xlsx"
        try:
            if not os.path.exists(excel_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Callbacks"
                ws.append(["Name", "Phone", "Preferred Time"])
                wb.save(excel_path)
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            ws.append([name, phone, preferred_time])
            wb.save(excel_path)
        except Exception as excel_error:
            pass  

        return jsonify({"success": True, "message": "Callback appointment booked successfully!", "details": callback})
    except Exception as e:
        return jsonify({"success": False, "message": f"Failed to book callback: {e}"}), 500

def get_related_questions(idx, questions, topn=3):
    scores = [(i, fuzz.token_sort_ratio(questions[idx], q)) for i, q in enumerate(questions) if i != idx]
    scores.sort(key=lambda x: x[1], reverse=True)
    related = [questions[i] for i, s in scores[:topn]]
    return related

def fetch_college_info(query):
    small_talk = {
        "hello": "Hello! 😊 How can I assist you with KIPM today?",
        "hi": "Hi there! 👋 How can I help you regarding KIPM?",
        "hey": "Hey! 🙌 What would you like to know about KIPM?",
        "good morning": "Good morning! ☀️ How can I help you with KIPM admissions?",
        "good afternoon": "Good afternoon! 🌞 How can I assist you today?",
        "good evening": "Good evening! 🌙 Need any information about KIPM?",
        "how are you": random.choice([
            "I'm just a bot, but I'm here and ready to help you with KIPM information! 🤖",
            "I'm doing great! Ready to help you with your KIPM queries! 😃"
        ]),
        "who are you": "I'm the KIPM Admission Chatbot 🤖, here to assist you with all your college queries.",
        "what can you do": "I can answer your questions about KIPM admissions, courses, fees, campus life, and more! 🎓",
        "tell me a joke": random.choice([
            "Why did the student eat his homework? Because the teacher said it was a piece of cake! 🍰😂",
            "Why was the math book sad? Because it had too many problems! 📚😅"
        ]),
        "you are smart": "Thank you! 🧠 I do my best to help you.",
        "you are helpful": "I'm glad to be of assistance! 😊",
        "i am confused": "No worries! 😌 Please tell me what you need help with regarding KIPM.",
        "i am lost": "I'm here to guide you. 🗺️ What would you like to know about KIPM?",
        "can you help me": "Absolutely! 🙋‍♂️ Please type your question about KIPM admissions or campus."
    }
    norm_query = query.lower().strip()
    if norm_query in small_talk and len(norm_query.split()) <= 5:
        return small_talk[norm_query]
    try:
        def normalize(text):
            return ' '.join(text.lower().strip().split()) if isinstance(text, str) else ''
        def get_synonyms(word):
            try:
                syns = set()
                for syn in wordnet.synsets(word):
                    for lemma in syn.lemmas():
                        name = lemma.name()
                        if isinstance(name, str):
                            syns.add(name.replace('_', ' '))
                return syns
            except Exception:
                # If wordnet is missing at runtime, attempt to download and retry once
                try:
                    nltk.download("wordnet")
                    nltk.download("omw-1.4")
                    syns = set()
                    for syn in wordnet.synsets(word):
                        for lemma in syn.lemmas():
                            name = lemma.name()
                            if isinstance(name, str):
                                syns.add(name.replace('_', ' '))
                    return syns
                except Exception:
                    return set()
        query_words = set(normalize(query).split())
        all_synonyms = set()
        for word in query_words:
            all_synonyms.update(get_synonyms(word))
        all_words = query_words | all_synonyms

        faqs = FAQS

        questions_en = [faq.get('question_en') for faq in faqs if faq.get('question_en')]
        answers_en = [faq.get('answer_en') for faq in faqs if faq.get('answer_en')]
        legacy_questions = [faq.get('question') for faq in faqs if faq.get('question')]
        legacy_answers = [faq.get('answer') for faq in faqs if faq.get('answer')]
        all_questions_en = questions_en + legacy_questions
        all_answers_en = answers_en + legacy_answers
        norm_questions_en = [normalize(q) for q in all_questions_en]
        norm_answers_en = [normalize(a) for a in all_answers_en]

        questions_hi = [faq.get('question_hi') for faq in faqs if faq.get('question_hi')]
        answers_hi = [faq.get('answer_hi') for faq in faqs if faq.get('answer_hi')]
        norm_questions_hi = [normalize(q) for q in questions_hi]
        norm_answers_hi = [normalize(a) for a in answers_hi]
        is_hindi = any(ord(char) >= 0x0900 and ord(char) <= 0x097F for char in query) or any(word in query_words for word in {'hai', 'kya', 'ka', 'mein', 'se', 'aur', 'ke', 'hain'})
        if is_hindi:
            questions = questions_hi
            answers = answers_hi
            norm_questions = norm_questions_hi
            norm_answers = norm_answers_hi
        else:
            questions = all_questions_en
            answers = all_answers_en
            norm_questions = norm_questions_en
            norm_answers = norm_answers_en

        threshold = 80
        best_idx = None
        best_score = 0
        for idx, q in enumerate(norm_questions):
            score = fuzz.token_sort_ratio(q, normalize(query))
            if score > best_score:
                best_score = score
                best_idx = idx
        if best_score >= threshold and best_idx is not None:
            answer = answers[best_idx] if best_idx < len(answers) else ""
            
            encouragements = [
                "Let me know if you have more questions! 🤗",
                "Feel free to ask anything else about KIPM! 💬",
                "I'm here if you need more info! 👍",
                "Hope this helps! 😊"
            ]
            answer = answer + " " + random.choice(["😊", "👍", "🎉", "🙌"])

            # related suggestions
            def get_related(idx, questions, topn=3):
                scores = [(i, fuzz.token_sort_ratio(questions[idx], q)) for i, q in enumerate(questions) if i != idx]
                scores.sort(key=lambda x: x[1], reverse=True)
                return [questions[i] for i, s in scores[:topn]]
            related = get_related(best_idx, questions) if questions else []
            if related:
                suggestion_text = " 💡 You may also ask: " + " | ".join(related)
                answer += suggestion_text

            answer += " " + random.choice(encouragements)
            session['unanswered_count'] = 0
            return answer
        else:

            OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")  
            if OPENAI_API_KEY:
                try:
                    headers = {
                        "Authorization": f"Bearer {OPENAI_API_KEY}",
                        "Content-Type": "application/json"
                    }
                    data = {
                        "model": "gpt-3.5-turbo",
                        "messages": [
                            {"role": "system", "content": "You are a helpful college admission assistant for KIPM."},
                            {"role": "user", "content": query}
                        ]
                    }
                    response = requests.post(
                        "https://api.openai.com/v1/chat/completions",
                        headers=headers,
                        json=data,
                        timeout=15
                    )
                    if response.status_code == 200:
                        llm_answer = response.json()["choices"][0]["message"]["content"]
                        return llm_answer + " 🤖"
                    else:
                        return (
                            "Sorry, i could't understand your Enquiry. If you any other doubt tell me i can help with us. Thank You 😔"
                            "Please visit the official KIPM website for more information:🌐"
                        )
                except Exception as llm_error:
                    return (
                        "Sorry, I couldn't find an exact answer to your question in our database. 😔"
                        "Please visit the official KIPM website for more information:🌐"
                    )
            else:
                return (
                    "Sorry, I couldn't find an exact answer to your question in our database. 😔"
                    "Please visit the official KIPM website for more information: "
                    "🌐"
                )
    except Exception as e:
        return f"Error reading FAQ data: {e} 😅"

@app.route("/")
def index():
    return render_template("index.html")

from flask import session

@app.route("/ask", methods=["POST"])
def ask():
    data = request.get_json()
    question = data.get("question", "").strip().lower()
    
    if session.get('eligibility_step') == 'ask_stream':
        stream = question
        session['eligibility_stream'] = stream
        session['eligibility_step'] = 'ask_marks'
        return jsonify({"answer": "Please enter your 12th percentage or marks (e.g. 78%)"})
    elif session.get('eligibility_step') == 'ask_marks':
        marks = question.replace('%','').replace('percent','').strip()
        try:
            marks = float(marks)
        except:
            return jsonify({"answer": "Please enter a valid number for your marks or percentage."})
        stream = session.get('eligibility_stream','')
        session.pop('eligibility_step', None)
        
        eligible = []
        if marks >= 45 and ('pcm' in stream or 'science' in stream or 'math' in stream):
            eligible.append('B.Tech (any branch)')
        if marks >= 40 and ('commerce' in stream or 'arts' in stream or 'any' in stream or 'all' in stream or 'science' in stream):
            eligible.append('BBA')
            eligible.append('BCA')
        if marks >= 50:
            eligible.append('MBA (if you are a graduate)')
        if eligible:
            return jsonify({"answer": "Based on your stream and marks, you are eligible for: " + ', '.join(eligible)})
        else:
            return jsonify({"answer": "Sorry, you do not meet the eligibility criteria for our main courses. Please contact the admission office for more options."})
    
    if 'eligibility' in question or 'eligible' in question or 'check admission' in question or 'can i apply' in question:
        session['eligibility_step'] = 'ask_stream'
        return jsonify({"answer": "To check your eligibility, please enter your 12th stream (e.g. PCM, Science, Commerce, Arts)"})

    answer = fetch_college_info(question)
    return jsonify({"answer": answer})

if __name__ == "__main__":
    app.run(debug=True, port=5050)


